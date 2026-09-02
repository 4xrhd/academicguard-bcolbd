"""
excel_gen.py — Excel report generation via OpenPyXL.
FR-RPT-02: Workbook with two sheets:
    - Summary: batch metadata + aggregate statistics
    - Analysis: one row per student with all scores, risk classification, and mark deductions
"""
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.engine.marking_calculator import calculate_marks

settings = get_settings()

# Risk level → Excel cell fill colour (hex)
_RISK_COLORS = {
    "high": "FF4C4C",    # Red
    "medium": "FFA500",  # Amber
    "low": "00B050"      # Green
}

_RISK_LIGHT_FILLS = {
    "high": "FEE2E2",
    "medium": "FEF3C7",
    "low": "DCFCE7"
}


async def generate(batch_id: uuid.UUID | str, db) -> Path:
    """
    Generate an Excel workbook for the batch.
    Returns the path to the generated .xlsx file.
    """
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload
    from app.db.models import Batch, Submission, RiskScore, AIDetectionResult

    b_uuid = uuid.UUID(str(batch_id))
    result = await db.execute(
        select(Batch)
        .where(Batch.id == b_uuid)
        .options(
            selectinload(Batch.instructor),
            selectinload(Batch.submissions).selectinload(Submission.risk_score),
            selectinload(Batch.submissions).selectinload(Submission.ai_result),
        )
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise ValueError(f"Batch {batch_id} not found.")

    export_dir = Path(settings.EXPORT_DIR) / str(batch.id)
    export_dir.mkdir(parents=True, exist_ok=True)
    output_path = export_dir / f"integrity_report_{batch.course_code}.xlsx"

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    if ws_summary is not None:
        ws_summary.title = "Summary"
        _populate_summary_sheet(ws_summary, batch)
    else:
        ws_summary = wb.create_sheet("Summary")
        _populate_summary_sheet(ws_summary, batch)

    # Sheet 2: Analysis
    ws_analysis = wb.create_sheet("Analysis")
    _populate_analysis_sheet(ws_analysis, batch)

    wb.save(str(output_path))
    return output_path


def _populate_summary_sheet(ws, batch) -> None:
    """Populate batch metadata and aggregate statistics."""
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    section_font = Font(name="Arial", size=11, bold=True, color="0F172A")
    bold_font = Font(name="Arial", size=10, bold=True)
    regular_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Title Banner
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "AcademicGuard | Classroom Integrity Audit Summary"
    title_cell.font = header_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Metadata
    submissions = batch.submissions or []
    total_subs = len(submissions)
    high_count = sum(1 for s in submissions if s.risk_score and s.risk_score.risk_level.lower() == "high")
    med_count = sum(1 for s in submissions if s.risk_score and s.risk_score.risk_level.lower() == "medium")
    low_count = total_subs - high_count - med_count

    avg_risk = sum(s.risk_score.weighted_score for s in submissions if s.risk_score) / total_subs if total_subs else 0.0
    avg_ai = sum(s.ai_result.final_ai_prob for s in submissions if s.ai_result) / total_subs if total_subs else 0.0

    meta_rows = [
        ("Batch Name", batch.name, "Total Submissions", total_subs),
        ("Course Code", batch.course_code, "Class Average Risk", f"{avg_risk*100:.1f}%"),
        ("Instructor", batch.instructor.full_name if batch.instructor else "N/A", "AI Detection Rate", f"{avg_ai*100:.1f}%"),
        ("Audit Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"), "High Risk Flagged", f"{high_count} ({(high_count/total_subs)*100:.1f}%)" if total_subs else "0"),
        ("Evaluation Status", (batch.status or "done").upper(), "Total Marks Possible", batch.total_marks or "N/A"),
    ]

    ws.append([])
    for row in meta_rows:
        ws.append([row[0], row[1], row[2], row[3]])
        curr_row = ws.max_row
        for col_idx in [1, 3]:
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = bold_font
            cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            cell.border = thin_border
        for col_idx in [2, 4]:
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border

    # Risk Distribution Table
    ws.append([])
    ws.append(["Risk Level Distribution", "", "", ""])
    ws.merge_cells(f"A{ws.max_row}:D{ws.max_row}")
    ws.cell(row=ws.max_row, column=1).font = section_font

    dist_headers = ["Classification", "Threshold", "Student Count", "Percentage"]
    ws.append(dist_headers)
    hdr_row = ws.max_row
    for c in range(1, 5):
        cell = ws.cell(row=hdr_row, column=c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    dist_rows = [
        ("LOW RISK", "< 25% Combined Risk", low_count, f"{(low_count/total_subs)*100:.1f}%" if total_subs else "0%", "DCFCE7"),
        ("MEDIUM RISK", "25% - 49% Combined Risk", med_count, f"{(med_count/total_subs)*100:.1f}%" if total_subs else "0%", "FEF3C7"),
        ("HIGH RISK", "≥ 50% Combined Risk", high_count, f"{(high_count/total_subs)*100:.1f}%" if total_subs else "0%", "FEE2E2"),
    ]

    for label, thresh, count, pct, fill_hex in dist_rows:
        ws.append([label, thresh, count, pct])
        r = ws.max_row
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            cell.border = thin_border
            if c in [1, 3, 4]:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)


def _populate_analysis_sheet(ws, batch) -> None:
    """Populate full student risk and grading analysis table."""
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Student ID", "Student Name", "Risk Level", "Weighted Score (%)",
        "AI Probability (%)", "Text Sim Max (%)", "Code Sim Max (%)",
        "Analysis Profile", "Total Deductions", "Final Marks Awarded"
    ]

    ws.append(headers)
    hdr_row = 1
    ws.row_dimensions[1].height = 26
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    submissions = batch.submissions or []
    for sub in submissions:
        rs = sub.risk_score
        ai = sub.ai_result

        weighted_score = (rs.weighted_score * 100) if rs else 0.0
        ai_prob = (ai.final_ai_prob * 100) if ai else ((rs.ai_prob * 100) if rs else 0.0)
        text_sim = (rs.text_sim_max * 100) if rs else 0.0
        code_sim = (rs.code_sim_max * 100) if rs and rs.code_sim_max is not None else None
        risk_lvl = (rs.risk_level if rs else "low").upper()
        profile = rs.weight_profile if rs else ("code_present" if sub.has_code else "theory_only")

        # Marks
        marks_obtained = sub.marks_obtained
        marks_breakdown = sub.marks_breakdown
        if (marks_obtained is None or marks_breakdown is None) and batch.total_marks and batch.marking_config and rs:
            marks_obtained, marks_breakdown = calculate_marks(
                total_marks=batch.total_marks,
                marking_config=batch.marking_config,
                ai_prob=rs.ai_prob,
                text_sim_max=rs.text_sim_max,
                code_sim_max=rs.code_sim_max,
                weighted_score=rs.weighted_score,
            )

        if marks_obtained is not None:
            tot_m = batch.total_marks if batch.total_marks is not None else 10.0
            deductions_val = max(0.0, tot_m - marks_obtained)
            deductions_str = f"-{deductions_val:.1f}" if deductions_val > 0.05 else "0.0"
            final_marks_str = f"{marks_obtained:.1f}"
        elif marks_breakdown and "total_deductions" in marks_breakdown:
            deductions_val = float(marks_breakdown["total_deductions"])
            deductions_str = f"-{deductions_val:.1f}" if deductions_val > 0.05 else "0.0"
            final_marks_str = "N/A"
        else:
            deductions_str = "N/A"
            final_marks_str = "N/A"

        row_data = [
            sub.student_id or "N/A",
            sub.student_name or "Unknown",
            risk_lvl,
            round(weighted_score, 1),
            round(ai_prob, 1),
            round(text_sim, 1),
            round(code_sim, 1) if code_sim is not None else "N/A",
            profile,
            deductions_str,
            final_marks_str
        ]
        ws.append(row_data)

        r = ws.max_row
        risk_key = risk_lvl.lower()
        fill_hex = _RISK_LIGHT_FILLS.get(risk_key, "FFFFFF")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=9.5)
            cell.border = thin_border
            if c == 3:  # Risk Level column
                cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
                cell.font = Font(name="Arial", size=9.5, bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif c in [1, 4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
